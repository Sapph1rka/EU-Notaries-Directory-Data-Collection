"""
EU Notaries Directory Data Collection Tool

This script scrapes notary information from the EU Notaries Directory website
(https://notaries-directory.eu/) and saves the data to an Excel file with
formatted styling.

The tool collects:
- Full Name
- First Name
- Email Address
- Country

Data is saved incrementally to prevent data loss and supports resuming
interrupted scraping sessions.
"""

import logging
import os
import threading
import time
from queue import Queue
from typing import Dict, List, Optional, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# Configuration constants
BASE_URL = "https://notaries-directory.eu/"
OUTPUT_FILENAME = "notaries_data.xlsx"
BATCH_SIZE = 7  # Number of notaries to process in parallel
PAGE_LOAD_DELAY = 1.5  # Seconds to wait for page load
THREAD_START_DELAY = 0.1  # Delay between thread starts
WAIT_TIMEOUT = 10  # Selenium wait timeout in seconds

# Excel formatting constants
HEADER_FILL_COLOR = "FFFF00"  # Yellow background for headers
COLUMN_WIDTHS = {
    "A": 60,  # Full Name
    "B": 40,  # First Name
    "C": 60,  # Email
    "D": 40,  # Country
}

# XPath selectors
XPATH_NOTARY_TITLE = '//*[@id="block-edn-theme-content"]/div/div/div/div[1]/div[1]/h1'
XPATH_COUNTRY = '//*[@id="block-edn-theme-content"]/div/div/div/div[1]/div[2]/div/span'
XPATH_EMAIL = '//div[contains(@class, "info-with-icon") and contains(@class, "link-redirect") and contains(@class, "mail")]/a'

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def setup_driver() -> webdriver.Chrome:
    """
    Initialize and configure Chrome WebDriver.

    Returns:
        webdriver.Chrome: Configured Chrome WebDriver instance
    """
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    driver = webdriver.Chrome(options=options)
    return driver


def get_notary_details(
    url: str, result_queue: Queue, full_name: str, index: int
) -> None:
    """
    Scrape detailed information for a single notary from their profile page.

    Args:
        url: URL of the notary's profile page
        result_queue: Queue to store the scraped data
        full_name: Full name of the notary (from listing page)
        index: Index position for sorting purposes
    """
    driver = setup_driver()
    try:
        driver.get(url)
        time.sleep(PAGE_LOAD_DELAY)

        # Wait for page to load
        WebDriverWait(driver, WAIT_TIMEOUT).until(
            EC.presence_of_element_located((By.XPATH, XPATH_NOTARY_TITLE))
        )

        # Extract country information
        try:
            country_element = driver.find_element(By.XPATH, XPATH_COUNTRY)
            country = country_element.text.split()[-1] if country_element else ""
        except NoSuchElementException:
            country = ""
            logger.warning(f"Country not found for {full_name}")

        # Extract email address
        try:
            email_element = driver.find_element(By.XPATH, XPATH_EMAIL)
            email = email_element.get_attribute("href").replace("mailto:", "")
        except NoSuchElementException:
            email = ""
            logger.debug(f"Email not found for {full_name}")

        # Prepare data record
        record = {
            "index": index,
            "Full Name": full_name,
            "First Name": full_name.split()[0] if full_name else "",
            "Email": email,
            "Country": country,
        }

        result_queue.put(record)
        logger.debug(f"Successfully scraped data for {full_name}")

    except TimeoutException:
        logger.error(f"Timeout while loading page for {full_name}: {url}")
        result_queue.put(None)
    except Exception as e:
        logger.error(f"Error scraping {full_name}: {str(e)}")
        result_queue.put(None)
    finally:
        driver.quit()


def format_excel(filename: str) -> None:
    """
    Apply formatting to the Excel file (headers, column widths, styling).

    Args:
        filename: Path to the Excel file to format
    """
    try:
        wb = load_workbook(filename)
        ws = wb.active

        # Configure header styling
        yellow_fill = PatternFill(
            start_color=HEADER_FILL_COLOR,
            end_color=HEADER_FILL_COLOR,
            fill_type="solid",
        )
        header_font = Font(bold=True)

        # Format header row (assuming 4 columns)
        num_columns = len(COLUMN_WIDTHS)
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = yellow_fill
            cell.font = header_font

        # Set column widths
        for col_letter, width in COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(filename)
        logger.debug(f"Applied formatting to {filename}")

    except Exception as e:
        logger.error(f"Error formatting Excel file {filename}: {str(e)}")


def save_to_excel(new_data: List[Dict[str, Any]], filename: str = OUTPUT_FILENAME) -> None:
    """
    Save notary data to Excel file, merging with existing data if present.

    Args:
        new_data: List of dictionaries containing notary information
        filename: Output Excel filename (default: notaries_data.xlsx)
    """
    if not new_data:
        logger.warning("No data to save")
        return

    try:
        # Create DataFrame from new data
        new_df = pd.DataFrame(new_data)

        # Merge with existing data if file exists
        if os.path.exists(filename):
            existing_df = pd.read_excel(filename)
            # Combine data and remove duplicates
            df = pd.concat([existing_df, new_df]).drop_duplicates()
            logger.info(f"Merged with existing data in {filename}")
        else:
            df = new_df
            logger.info(f"Creating new file: {filename}")

        # Sort by index and remove index column
        df = df.sort_values("index")
        df = df.drop("index", axis=1)

        # Save to Excel
        df.to_excel(filename, index=False)

        # Apply formatting
        format_excel(filename)

        logger.info(f"Data saved to {filename}. Total records: {len(df)}")

    except Exception as e:
        logger.error(f"Error saving data to {filename}: {str(e)}")
        raise


def process_notaries(
    elements: List[Any], current_page: int
) -> List[Dict[str, Any]]:
    """
    Process a batch of notary listing elements in parallel using threads.

    Args:
        elements: List of Selenium WebElement objects representing notary listings
        current_page: Current page number (for logging purposes)

    Returns:
        List of dictionaries containing scraped notary data
    """
    result_queue = Queue()
    threads = []
    details = []

    # Create threads for parallel processing
    for index, element in enumerate(elements):
        try:
            # Extract full name and URL from listing element
            full_name = element.find_element(By.XPATH, ".//h3").text
            href = element.find_element(By.TAG_NAME, "a").get_attribute("href")

            if href:
                thread = threading.Thread(
                    target=get_notary_details,
                    args=(href, result_queue, full_name, index),
                )
                threads.append(thread)
                thread.start()
                time.sleep(THREAD_START_DELAY)  # Small delay between thread starts
            else:
                logger.warning(f"No URL found for {full_name}")

        except Exception as e:
            logger.error(f"Error extracting data from listing element: {str(e)}")
            continue

    # Wait for all threads to complete
    for thread in threads:
        thread.join()

    # Collect results from queue
    while not result_queue.empty():
        detail = result_queue.get()
        if detail:
            details.append(detail)

    # Save data incrementally
    if details:
        save_to_excel(details)

    return details


def main() -> None:
    """
    Main function to orchestrate the web scraping process.

    Iterates through all pages of the notaries directory, extracts listing
    information, and scrapes detailed data for each notary.
    """
    main_driver = setup_driver()
    page = 0
    all_data = []

    try:
        while True:
            url = f"{BASE_URL}en/search?page={page}"
            logger.info(f"Processing page {page}...")

            main_driver.get(url)
            time.sleep(PAGE_LOAD_DELAY)

            # Wait for list elements to load
            try:
                WebDriverWait(main_driver, WAIT_TIMEOUT).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "list-element"))
                )
            except TimeoutException:
                logger.info("No more pages to scrape")
                break

            # Get all notary listing elements
            list_elements = main_driver.find_elements(By.CLASS_NAME, "list-element")
            logger.info(f"Found {len(list_elements)} notaries on page {page}")

            if not list_elements:
                logger.info("No more notaries found")
                break

            # Process elements in batches to avoid overwhelming the server
            for i in range(0, len(list_elements), BATCH_SIZE):
                current_elements = list_elements[i : i + BATCH_SIZE]
                end_index = min(i + BATCH_SIZE, len(list_elements))
                logger.info(
                    f"Processing elements {i+1} to {end_index} of {len(list_elements)}"
                )
                details = process_notaries(current_elements, page)
                all_data.extend(details)

            # Check if next page exists
            next_page = page + 1
            next_url = f"{BASE_URL}en/search?page={next_page}"
            main_driver.get(next_url)
            time.sleep(PAGE_LOAD_DELAY)

            try:
                next_list_elements = main_driver.find_elements(
                    By.CLASS_NAME, "list-element"
                )
                if not next_list_elements:
                    logger.info("No more notaries found on next page")
                    break
                page = next_page
            except Exception as e:
                logger.error(f"Error checking next page: {str(e)}")
                break

    except KeyboardInterrupt:
        logger.info("Scraping interrupted by user")
    except Exception as e:
        logger.error(f"Unexpected error in main loop: {str(e)}")
    finally:
        main_driver.quit()
        logger.info("Scraping completed. Driver closed.")
        logger.info(f"Total records collected: {len(all_data)}")


if __name__ == "__main__":
    main()
